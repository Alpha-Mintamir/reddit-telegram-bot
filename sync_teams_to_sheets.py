import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

from app.config import BotConfig
from app.integrations.google_sheets_client import GoogleSheetsClient
from openpyxl import load_workbook
import gspread

# Load config
config = BotConfig.from_env(require_reddit=False)
sheets = GoogleSheetsClient(config)

# Read Teams data from Excel
excel_path = Path(__file__).parent.parent / "data" / "outputs" / "reddit_team_assignments.xlsx"
wb = load_workbook(excel_path)
ws = wb['Teams']

# Get headers
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print(f'Headers: {headers}')

# Get all rows
rows_data = []
for r in range(2, ws.max_row + 1):
    row = {}
    for c, header in enumerate(headers, start=1):
        value = ws.cell(r, c).value
        row[header] = str(value).strip() if value is not None else ''
    rows_data.append(row)

print(f'Found {len(rows_data)} team member rows')

# Ensure Teams tab exists with headers
teams_ws = sheets.get_or_create_worksheet(config.teams_tab_name, headers=headers)

# Clear existing data (keep header row)
existing_rows = teams_ws.get_all_values()
if len(existing_rows) > 1:
    # Delete all rows except header
    teams_ws.delete_rows(2, len(existing_rows))
    print(f'Cleared {len(existing_rows) - 1} existing data rows')

# Write new rows
for row_data in rows_data:
    row_values = [row_data.get(h, '') for h in headers]
    teams_ws.append_row(row_values)
    print(f'  Added: {row_data.get("member_name")} (team {row_data.get("team_id")}, @{row_data.get("telegram_user_id", "").lstrip("@")})')

print(f'\nSuccessfully synced {len(rows_data)} team members to Google Sheets Teams tab')
